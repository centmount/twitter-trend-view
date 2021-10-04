#!/usr/bin/env python3

"""
日時を入れたファイル名を指定した「空のExcelシート」を作成してストレージにアップロードします。
[Google Cloud Platform]の[Cloud Functions]にデプロイして、[Cloud Scheduler]で定期実行します。
[Storge]にExcelファイルを保存します。
"""

#必要なモジュールのインポート
from datetime import datetime, timedelta
import pytz
import openpyxl
from google.cloud import storage

# 変数
japan_time = datetime.now() + timedelta(hours=9)
wb = openpyxl.Workbook()
wb.create_sheet()
ws = wb.worksheets[0]
ws_new = wb.worksheets[1] 


# 空のExcelシートの作成
def create_file():
    for i in range(50):
        ws.cell(row=i+2, column = 1).value = i+1
        ws_new.cell(row=i+2, column = 1).value = i+1
    wb.save(f'/tmp/trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx')


# Storageへのアップロード処理
def upload_blob(bucket_name, source_file_name, destination_blob_name):
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)
    blob.upload_from_filename(source_file_name)  

# main関数を実行(GCPのCloud Functions用にevent,contextが引数に)
# トリガーをpub/subに設定、ファイルダウンロードしてデータ追加してアップロード
def main(event, context):
    create_file()
    upload_blob('ストレージのバケット名を指定', f'/tmp/trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx', f'trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx')
      