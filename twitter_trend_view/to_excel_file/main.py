#!/usr/bin/env python3

"""
Twitter APIからトレンドデータを定期的にexcelファイルに保存します。
トップ50のランキングのデータを10分ごとに取得して、日時順に追加保存します。
[Google Cloud Platform]の[Cloud Functions]にデプロイして、[Cloud Scheduler]で定期実行します。
[Storge]にExcelファイルを更新して保存します。
"""

# Twitterトレンドを取得してExcel保存する
#必要なモジュールのインポート
import tweepy
import time
from datetime import datetime, timedelta
import pytz
import openpyxl
from openpyxl.styles import PatternFill
from google.cloud import storage
import os

# TwitterAPIの各種ツイッターのキーを入力（必要に応じてシークレットファイルなどを使用してください）
def input_twitter_info():
    global CONSUMER_KEY
    global CONSUMER_SECRET
    global ACCESS_KEY
    global ACCESS_SECRET
    while True:
        CONSUMER_KEY = input('TwitterAPIのCONSUMER_KEYを入力して下さい: ') # TwitterAPIのキーが必要
        CONSUMER_SECRET = input('TwitterAPIのCONSUMER_SECRETを入力して下さい: ')
        ACCESS_KEY = input('TwitterAPIのACCESS_KEYを入力して下さい: ')
        ACCESS_SECRET = input('TwitterAPIのACCESS_SECRETを入力して下さい: ')
        ok = input('入力完了しましたか？(完了の場合 OK と入力): ')
        if ok == 'OK':
            break
    return CONSUMER_KEY, CONSUMER_SECRET, ACCESS_KEY, ACCESS_SECRET

# OAuth認証
def authTwitter():
    auth = tweepy.OAuthHandler(CONSUMER_KEY, CONSUMER_SECRET)
    auth.set_access_token(ACCESS_KEY, ACCESS_SECRET)

    #APIインスタンスを作成、レート制限が補充されるまで待機、残り時間を知らせる
    api = tweepy.API(auth, wait_on_rate_limit = True, wait_on_rate_limit_notify=True)
    return(api)

#日本のトレンドランキングを取得

woeid = {"日本": 23424856}
api = authTwitter()
japan_time = datetime.now() + timedelta(hours=9)

wb = openpyxl.Workbook()
wb.create_sheet()
ws = wb.worksheets[0]
ws_new = wb.worksheets[1] 


def create_file():
    for i in range(50):
        ws.cell(row=i+2, column = 1).value = i+1
        ws_new.cell(row=i+2, column = 1).value = i+1
    wb.save(f'/tmp/trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx')


def trend():
    trend_data = []
    for area, wid in woeid.items():
        japan_time = datetime.now() + timedelta(hours=9)
        print(japan_time)
        trends = api.trends_place(wid)[0]
        for i, content in enumerate(trends["trends"]):
            [a, b] = [i+1, content["name"]]
            print(a, b)
            trend_data.append(b)
    print(trend_data)

    wb = openpyxl.load_workbook(f'/tmp/trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx')
    ws = wb.worksheets[0]
    ws.title = "AllData"
    ws_new = wb.worksheets[1]
    ws_new.title = "NewData"
    wc = ws.max_column + 1

    ws.cell(row=1, column = wc).value = japan_time
    for i, data in enumerate(trend_data):
        row = i + 2
        c = ws.cell(row=row, column = wc)
        c.value = data
        for j in range(2, ws.max_column):
            for i in range(2, ws.max_row + 1):
                if c.value == ws.cell(row=i, column=j).value:
                    c.fill = PatternFill(fgColor="40E0D0", bgColor="40E0D0", fill_type = "solid")
      
    #セルの幅を自動設定する
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 6) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width
  
    #コピー元の列を読み取って、書き込む
    if ws.max_column >= 6:
        for m in range(ws.max_column - 4 , ws.max_column + 1):
            for n in range(1, ws.max_row + 1):
                ws_new.cell(row = n, column = m - ws.max_column + 6).value = ws.cell(row = n, column = m).value
                #if ws.cell(row = n, column = m).has_style:
                ws_new.cell(row = n, column = m - ws.max_column + 6)._style = ws.cell(row = n, column = m)._style
    else:
        pass

    #セルの幅を自動設定する
    for col in ws_new.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 6) * 1.2
        ws_new.column_dimensions[col[0].column_letter].width = adjusted_width

    wb.save(f'/tmp/trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx')


# Storageへのアップロード処理
def upload_blob(bucket_name, source_file_name, destination_blob_name):
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)
    blob.upload_from_filename(source_file_name)  

# Storageからのダウンロード処理
def download_blob(bucket_name, source_blob_name, destination_file_name):
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(source_blob_name)
    blob.download_to_filename(destination_file_name)

# GCPの[Cloud Functions]にデプロイ用
# トリガーをpub/subに設定、ファイルダウンロードしてデータ追加してアップロード
def main(event, context):
    input_twitter_info()
    download_blob('ストレージのバケット名を指定', f'trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx', f'/tmp/trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx')
    trend()
    upload_blob('ストレージのバケット名を指定', f'/tmp/trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx', f'trend_data{japan_time.strftime("%Y年%m月%d日")}.xlsx')

    